import copy
import xlsxwriter
from openpyxl import load_workbook


SMART_EXCEL_CONFIG = {
    'sheet_names': ['Sheet1', '_data', '_meta'],
    'dump_date_cell_position': 'B1',
    'header_row_cell_position': 'B2'
}


class SmartExcel():
    """
    The SmartExcel class is responsible for two things:
        * Dump data from a definition to a xlsx file.
        * Parse a xlsx and retrieve data from a definition.

    ATTRIBUTES:
    header_row: number of rows used as Header.
    max_row: number of rows where validation is applied.
    meta_worksheet_name: Name of the 'meta' worksheet (used to store timestamp and settings).
    data_worksheet_name = Name of the 'data' worksheet (used to store list options)
    READMODE: If this is set to True, only parsing is available.
    WRITEMODE: If this is set to true, only dumping is available.
    """  # noqa
    header_row = 1
    max_row = 100
    meta_worksheet_name = '_meta'
    data_worksheet_name = '_data'
    READMODE = False
    WRITEMODE = False

    def __init__(
        self,
            definition=None,
            data=None,
            path=None,
            output='template.xlsx'):
        """
        Init a new instance of the SmartExcel class.

        :param definition: A definition of the xlsx (headers, columns, validations).
        :type definition: list

        :param data: An helper class to retrieve data based on the definition.
        :type data: object

        :param path: The path to a xlsx file. Only in READMODE.
        :type path: string

        :param output: The output of a xlsx file. Only in WRITEMODE.
        :type output: str or io.BytesIO()
        """  # noqa

        assert definition and data

        self.sheets = {}
        # self.columns = []
        self.formats = {}
        self.validations = {}
        self.groups = {}

        self.data = data

        if path:
            self.READMODE = True
            self.init_read_mode(definition, path)
        else:
            self.WRITEMODE = True
            self.init_write_mode(definition, output)


    def init_read_mode(self, definition, path):
        """
        Init in READMODE.
        """
        self.parse_definition(definition)

        self.workbook = load_workbook(path)

        self.meta_config = check_meta_config(self.workbook)

        header = [c['name'] for c in self.columns]

        check_header(
            self.workbook['Sheet1'],
            header,
            self.meta_config['header_row'])

    def init_write_mode(self, definition, output):
        """
        Init in WRITEMODE.
        """

        self.output = output
        self.workbook = xlsxwriter.Workbook(self.output)
        # self.main_ws = self.workbook.add_worksheet()

        self.data_ws = self.workbook.add_worksheet(self.data_worksheet_name)
        self.data_ws.protect()

        self.meta_ws = self.workbook.add_worksheet(self.meta_worksheet_name)
        self.meta_ws.protect()

        self.parse_definition(definition)

    def parse(self):
        """
        Parse a xlsx file according to the definition and returns an list of dict (rows).
        """  # noqa
        assert self.READMODE

        self.parsed_data = []
        for row_index, row in enumerate(self.workbook['Sheet1'].values):
            parsed_row = {}
            if row_index < self.meta_config['header_row']:
                continue
            for col_index, col in enumerate(self.columns):
                if col['index'] > 0 and col['key'].find('--') == -1:
                    new_key = '{key}--{index}'.format(
                        key=col['key'],
                        index=col['index'])
                    col['key'] = new_key

                col_index_base_1 = col_index + 1
                row_index_base_1 = row_index + 1
                value = self.workbook['Sheet1'].cell(
                    column=col_index_base_1,
                    row=row_index_base_1).value

                parsed_row[col['key']] = value

            self.parsed_data.append(parsed_row)
        return self.parsed_data

    def dump(self):
        """
        Dump data into a xlsx file according to the definition.
        """
        assert self.WRITEMODE

        # self.build_top_header()
        self.build_meta()
        self.build_data()

        for sheet_key, sheet_data in self.sheets.items():
            fd_current_sheet = self.workbook.add_worksheet(sheet_data['name'])

            max_stack = len(sheet_data['components'])
            current_component = None
            next_available = {
                'row': 1,
                'col': 0
            }

            for x in range(0, max_stack):
                for y in range(0, max_stack):
                    for component in sheet_data['components']:
                        if component['stack']['x'] == x and component['stack']['y'] == y:

                            current_component = component
                            if 'rows' in component:

                                for index, row in enumerate(component['rows']):
                                    n_col = next_letter(next_available['col'])
                                    n_row = next_available['row'] + index + 1

                                    cell_pos = f"{n_col}{n_row}"

                                    if row['name']:
                                        value = [
                                            f"{row['name']}",
                                            self.get_value(self.data, f"write_{row['key']}", component['payload'][0], {})
                                        ]
                                    else:
                                        value = [
                                            self.get_value(self.data, f"write_{row['key']}", component['payload'][0], {})
                                        ]
                                    fd_current_sheet.write_row(cell_pos, value)
                                next_available['col'] += len(value) + 2
                                next_available['row'] += len(component['rows']) + 2

                            elif 'columns' in component:
                                for column in component['columns']:
                                    self.write_header(fd_current_sheet, column, next_available)
                                    self.set_validations(fd_current_sheet, column)

                                    values = self.get_values_for_column(column, component['payload'])

                                    self.set_column_width(fd_current_sheet, column['letter'], values)

                                    cell_format = self.get_column_format(column)

                                    for index, value in enumerate(values):
                                        cell_pos = f'{column["letter"]}{index + next_available["row"] + 1 + self.header_row}'
                                        fd_current_sheet.write(cell_pos, value, cell_format)

                                next_available['row'] += len(values) + 1 + 2 # self.SPACE_COMPONENT

        self.workbook.close()

    def add_sheet(self, definition, index=0):
        # try:
        #     import pdb ; pdb.set_trace()
        #     sheet_name = self.get_value(
        #         self.data,
        #         f"get_sheet_name_for_{definition['def_sheet_name']['func']}",
        #         self.data.results[definition['def_sheet_name']['payload']],
        #         {})
        # except KeyError as error:
            # import pdb ; pdb.set_trace()
            # sheet_name = definition.get('sheet_name_func', f'default-{index}')
        try:
            sheet_name = definition['sheet_name']
        except KeyError:
            sheet_name = getattr(
                self.data,
                f"get_sheet_name_for_{definition['sheet_name_func']}")()

        print(f'Add {sheet_name}')

        try:
            sheet_key = f"{definition['key']}-{index}"
        except KeyError:
            sheet_key = f'{sheet_name}-{index}'

        self.sheets[sheet_key] = {
            'name': sheet_name,
            'components': []
        }

        self.parse_components(definition['components'], **{
            'sheet_key': sheet_key
        })

    def parse_components(self, components, **kwargs):
        for component in components:
            kwargs.update(component)
            try:
                getattr(self, component['func'])(**kwargs)
            except AttributeError as e:
                raise Exception(f'{e} not supported')
            if 'recursive' in component:
                for index, instance in enumerate(self.data.results[component['payload']]):

                    payload = getattr(self.data, f"get_payload_{component['recursive']['payload_func']}")(
                        instance=instance,
                        foreign_key=component['recursive']['foreign_key']
                    )
                    self.data.results[component['recursive']['payload_func']] = payload

                    for rec_component in component['recursive']['components']:
                        rec_component['payload'] = component['recursive']['payload_func']

                    sheet_name = getattr(
                        self.data,
                        f"get_sheet_name_for_{component['recursive']['sheet_name_func']}")(
                            instance
                        )

                    self.add_sheet(
                        definition={
                            'sheet_name': sheet_name,
                            'components': component['recursive']['components']
                        },
                        index=index)



    def call_ext_func(self, func, *args, **kwargs):
        if func == 'add_sheet':
            self.add_sheet(*args, **kwargs)
        elif func == 'add_format':
            self.add_format(args[0]['kwargs'])



    def parse_definition(self, definition):
        for elem in definition:
            self.call_ext_func(elem['func'], elem)

    def add_format(self, cell_format):
        self.formats[cell_format['key']] = self.workbook.add_format(
            cell_format['format'])

        if 'num_format' in cell_format:
            self.formats[cell_format['key']].set_num_format(
                cell_format['num_format'])


    def add_group_row(self, **kwargs):
            # and if 'header' in deef['components']:

        parsed_rows = []
        for row in kwargs['rows']:
            tmp = copy.deepcopy(row)

            tmp.update({
                'letter': 'A'
            })

            parsed_rows.append(tmp)

        sheet_key = kwargs['sheet_key']
        self.sheets[sheet_key]['components'].append({
            'payload': self.data.results[kwargs['payload']],
            'name': kwargs['name'],
            'rows': parsed_rows,
            'stack': kwargs['stack']
        })

    def add_group_column(self, **kwargs):
        try:
            group_name = kwargs['group_name']
        except:
            group_name = None


        try:
            if 'repeat_func' in kwargs:
                repeat = getattr(
                    self.data,
                    'write_{key}'.format(
                        key=kwargs['repeat_func']))()
            else:
                repeat = kwargs['repeat']
        except Exception:
            repeat = 1

        parsed_columns = []
        for index in range(0, repeat):
            for column in kwargs['columns']:
                tmp = copy.deepcopy(column)

                if 'name_func' in tmp:
                    tmp['name'] = self.get_value(
                        self.data,
                        'write_{key}'.format(key=tmp['name_func']),
                        None, {'index': index})
                    del tmp['name_func']

                if repeat > 1:
                    name = f'{tmp["name"]} - {index + 1}'
                else:
                    name = tmp['name']

                tmp.update({
                    'name': name,
                    'letter': next_letter(len(parsed_columns)),
                    'index': index
                })

                if group_name in self.groups:
                    self.groups[group_name]['end'] = tmp['letter']
                else:
                    self.groups[group_name] = {
                        'start': tmp['letter'],
                        'end': tmp['letter']
                    }

                parsed_columns.append(tmp)

        sheet_key = kwargs['sheet_key']

        self.sheets[sheet_key]['components'].append({
            'payload': self.data.results[kwargs['payload']],
            'name': kwargs['name'],
            'columns': parsed_columns,
            'stack': kwargs['stack']
        })

    def write_header(self, sheet, column, next_available):
        col = column["letter"]
        # next_letter(next_available['col'])
        #

        row = self.header_row + next_available['row']

        cell_pos = f'{col}{row}'

        if 'required' in column:
            cell_format = 'header_required'
        else:
            cell_format = 'header'

        sheet.write(cell_pos, column['name'], self.get_format(cell_format))

    def set_list_source_func(self, sheet, cell_range, column):
        if 'list_source_func' in column['validations']:
            sheet.data_validation(cell_range, {
                'validate': 'list',
                'source': f'={self.validations[column["key"]]["meta_source"]}'
            })

    def column_cell_range(self, column):
        return '{start_letter}{start_pos}:{end_letter}{end_pos}'.format(
            start_letter=column["letter"],
            start_pos=self.header_row + 1,
            end_letter=column["letter"],
            end_pos=self.max_row
        )
        # return f'{}{}:{column["letter"]}{self.max_row}'

    def set_validations(self, sheet, column):
        if column["key"] in self.validations:
            cell_range = self.column_cell_range(column)

            self.set_list_source_func(sheet, cell_range, column)
            self.set_excel_validations(sheet, cell_range, column)

    def set_excel_validations(self, sheet, cell_range, column):
        if 'excel' in column['validations']:
            sheet.data_validation(
                cell_range,
                column['validations']['excel'])

    def build_meta(self):
        from datetime import datetime
        now = datetime.now().strftime('%Y-%m-%d')
        self.meta_ws.write_row('A1', ['dump_date', now])
        self.meta_ws.write_row('A2', ['header_rows', self.header_row])

    def build_data(self):
        for sheet_key, sheet_data in self.sheets.items():
            for component in sheet_data['components']:
                if 'rows' in component:
                    pass
                if 'columns' in component:
                    for column in component['columns']:
                        if column['key'] not in self.validations:
                            tmp = {
                                'row': len(self.validations) + 1
                            }
                            tmp.update(column['validations'])

                            if 'list_source_func' in column['validations']\
                                and column['key'] not in self.validations:
                                list_source = getattr(
                                    self.data,
                                    column['validations']['list_source_func'])()

                                tmp.update({
                                    'meta_source': f'={self.data_worksheet_name}!$A${tmp["row"]}:${next_letter(len(list_source) - 1)}${tmp["row"]}'  # noqa
                                })

                                self.data_ws.write_row(
                                    f"A{tmp['row']}",
                                    list_source)

                            self.validations[column['key']] = tmp

    def get_format(self, format_name):
        if format_name in self.formats:
            return self.formats[format_name]
        else:
            None

    def get_column_format(self, column):
        if 'format' in column:
            return self.formats[column['format']]
        else:
            return None

    def build_top_header(self):
        if self.groups:
            self.header_row += 1
            for group_name, position in self.groups.items():
                if group_name is None:
                    continue

                if position['start'] == position['end']:
                    self.main_ws.set_column(
                        f"{position['start']}:{position['start']}",
                        len(group_name))

                    self.main_ws.write(
                        f"{position['start']}1",
                        group_name,
                        self.get_format('top_header'))

                else:
                    cell_range = f"{position['start']}1:{position['end']}1"
                    self.main_ws.merge_range(
                        cell_range,
                        group_name,
                        self.get_format('top_header'))

        self.main_ws.freeze_panes(1, 0)
        self.main_ws.freeze_panes(2, 0)

    def set_column_width(self, fd_current_sheet, column_letter, values):
        try:
            width = len(max(values, key=len))
            if width < 10:
                width = 10
        except (TypeError, ValueError):
            width = 10

        fd_current_sheet.set_column(
            f'{column_letter}:{column_letter}',
            width)

    def get_meta(self, klass, func, meta, kwargs):
        if func not in dir(klass):
            raise Exception(f'method \'{func}\' not present in SmartExcelData class')  # noqa
        try:
            meta = getattr(klass, func)(meta, kwargs)
        except IndexError:
            meta = None

        return meta

    def get_value(self, klass, func, obj, kwargs):
        return self.get_meta(klass, func, obj, kwargs)

    def get_values_for_column(self, column, results):
        return [
            self.get_value(
                self.data,
                'write_{key}'.format(key=column['key']),
                obj,
                {'index': index})
            for index, obj in enumerate(results)  # self.data.results
        ]


def check_sheet_names(sheet_names):
    if sheet_names != SMART_EXCEL_CONFIG['sheet_names']:
        raise Exception("'Sheet1', '_meta', '_data' sheets must be present.")


def check_dump_date(meta_ws):
    dump_date = meta_ws[SMART_EXCEL_CONFIG['dump_date_cell_position']]
    if dump_date is None:
        raise Exception("A dump date must be present.")
    return dump_date.value


def check_header_row(meta_ws):
    header_row = meta_ws[SMART_EXCEL_CONFIG['header_row_cell_position']]
    if header_row is None:
        raise Exception("config header_row must be present.")
    return header_row.value


def check_header(sheet, definition, header_row):
    header_row_base_0 = header_row - 1
    # sheet.values is a generator, so an iterator. I should use itertools
    for index, row in enumerate(sheet.values):
        if index == header_row_base_0:
            if tuple(definition) != row:
                raise Exception("Header definitions do not match.")


def check_meta_config(wb):
    check_sheet_names(wb.sheetnames)

    return {
        'dump_date': check_dump_date(wb['_meta']),
        'header_row': check_header_row(wb['_meta'])
    }



A, Z = 65, 90
TOTAL = 26


def next_letter(length):
    """
    Get the next excel column available.
    e.g: 'A', 'AA', 'BA'

    length: of the columns list
    returns
    """
    pos = int(length / TOTAL)

    if length >= TOTAL:
        next_length = length - (TOTAL * pos)
        return f'{next_letter(pos - 1)}{next_letter(next_length)}'
    else:
        char = length + A
    return chr(char)
